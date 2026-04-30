import time
import requests
import openpyxl
from openpyxl.styles import PatternFill, Font
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager

BASE_URL   = "http://localhost:3000"
EXCEL_PATH = "dados.xlsx"
TIMEOUT    = 10

# cores pra colorir as células no excel
FILL_OK   = PatternFill("solid", fgColor="D1FAE5")
FILL_ERRO = PatternFill("solid", fgColor="FEE2E2")
FONT_OK   = Font(color="065F46", bold=True)
FONT_ERRO = Font(color="991B1B", bold=True)


def iniciar_driver():
    opts = webdriver.ChromeOptions()
    opts.add_argument("--start-maximized")
    service = Service(ChromeDriverManager().install())
    driver  = webdriver.Chrome(service=service, options=opts)
    driver.get(BASE_URL)
    return driver


def esperar(driver, by, valor, timeout=TIMEOUT):
    return WebDriverWait(driver, timeout).until(
        EC.presence_of_element_located((by, valor))
    )


def esperar_clicavel(driver, by, valor, timeout=TIMEOUT):
    return WebDriverWait(driver, timeout).until(
        EC.element_to_be_clickable((by, valor))
    )


def clicar_aba(driver, aba):
    esperar_clicavel(driver, By.CSS_SELECTOR, "[data-tab='%s']" % aba).click()
    time.sleep(0.5)


def preencher(el, valor):
    el.clear()
    el.send_keys(str(valor))


def buscar_api(driver, endpoint):
    # usa XHR síncrono pra não complicar com async
    return driver.execute_script(
        "var xhr = new XMLHttpRequest();"
        "xhr.open('GET', '%s', false);"
        "xhr.send();"
        "return JSON.parse(xhr.responseText);" % endpoint
    )


def ler_aba(wb, nome_aba):
    ws = wb[nome_aba]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    linhas = []
    for row in ws.iter_rows(min_row=2, values_only=False):
        if not any(c.value for c in row):
            continue
        linhas.append({headers[i]: row[i] for i in range(len(headers))})
    return headers, linhas


def marcar(linha, col, valor, ok):
    cell = linha[col]
    cell.value = valor
    cell.fill  = FILL_OK if ok else FILL_ERRO
    cell.font  = FONT_OK if ok else FONT_ERRO


def val(linha, col):
    cell = linha.get(col)
    if cell is None or cell.value is None:
        return ""
    return str(cell.value).strip()


def cadastrar_clientes(driver, wb):
    print("\n-- Clientes --")
    _, linhas = ler_aba(wb, "Clientes")
    clicar_aba(driver, "clientes")

    for i, linha in enumerate(linhas):
        email = val(linha, "email")

        # pula linhas que já foram cadastradas com sucesso
        if val(linha, "status") == "sucesso":
            print("  [%d] %s — pulando" % (i + 1, email))
            continue

        try:
            esperar_clicavel(driver, By.ID, "btn-limpar-cliente").click()
            time.sleep(0.3)

            preencher(driver.find_element(By.ID, "nome"),      val(linha, "nome"))
            preencher(driver.find_element(By.ID, "sobrenome"), val(linha, "sobrenome"))

            if val(linha, "data_nascimento"):
                cell_val = linha["data_nascimento"].value
                if isinstance(cell_val, datetime):
                    data_fmt = cell_val.strftime("%Y-%m-%d")
                else:
                    raw = str(cell_val).strip()
                    data_fmt = raw
                    # tenta alguns formatos comuns de data
                    for fmt in ("%d/%m/%Y", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%m/%d/%Y"):
                        try:
                            data_fmt = datetime.strptime(raw, fmt).strftime("%Y-%m-%d")
                            break
                        except ValueError:
                            continue

                # o input type=date no Windows pode reordenar os valores,
                # manda dia, mês e ano separados pra garantir
                campo_data = driver.find_element(By.ID, "data_nascimento")
                campo_data.click()
                partes = data_fmt.split("-")
                ano, mes, dia = partes[0], partes[1], partes[2]
                campo_data.send_keys(dia + mes + ano)

            if val(linha, "telefone"):
                preencher(driver.find_element(By.ID, "telefone"), val(linha, "telefone"))

            preencher(driver.find_element(By.ID, "email"), email)

            campo_cep = driver.find_element(By.ID, "cep")
            preencher(campo_cep, val(linha, "cep"))
            campo_cep.send_keys(Keys.RETURN)

            # espera o campo cidade ser preenchido pela BrasilAPI
            WebDriverWait(driver, 8).until(
                lambda d: d.find_element(By.ID, "cidade").get_attribute("value") != ""
            )
            time.sleep(0.3)

            esperar_clicavel(driver, By.ID, "btn-salvar-cliente").click()
            WebDriverWait(driver, TIMEOUT).until(
                EC.text_to_be_present_in_element((By.ID, "msg-cliente"), "cadastrado")
            )

            marcar(linha, "status", "sucesso", ok=True)
            marcar(linha, "erro",   "",        ok=True)
            print("  [%d] %s — OK" % (i + 1, email))

        except Exception as e:
            msg = repr(e).split("(")[0] + ": " + str(e).split("\n")[0][:150]
            marcar(linha, "status", "erro", ok=False)
            marcar(linha, "erro",   msg,    ok=False)
            print("  [%d] %s — ERRO: %s" % (i + 1, email, msg))
            try:
                esperar_clicavel(driver, By.ID, "btn-limpar-cliente").click()
            except Exception:
                pass

        finally:
            wb.save(EXCEL_PATH)
            time.sleep(0.5)


def cadastrar_produtos(driver, wb):
    print("\n-- Produtos --")
    _, linhas = ler_aba(wb, "Produtos")
    clicar_aba(driver, "produtos")

    for i, linha in enumerate(linhas):
        codigo = val(linha, "codigo")

        if val(linha, "status") == "sucesso":
            print("  [%d] %s — pulando" % (i + 1, codigo))
            continue

        try:
            esperar_clicavel(driver, By.ID, "btn-limpar-produto").click()
            time.sleep(0.3)

            preencher(driver.find_element(By.ID, "prod-codigo"), codigo)
            preencher(driver.find_element(By.ID, "prod-nome"),   val(linha, "nome"))
            preencher(driver.find_element(By.ID, "prod-valor"),  val(linha, "valor"))

            if val(linha, "descricao"):
                preencher(driver.find_element(By.ID, "prod-descricao"), val(linha, "descricao"))

            driver.find_element(By.CSS_SELECTOR, "#form-produto button[type='submit']").click()
            WebDriverWait(driver, TIMEOUT).until(
                EC.text_to_be_present_in_element((By.ID, "msg-produto"), "cadastrado")
            )

            marcar(linha, "status", "sucesso", ok=True)
            marcar(linha, "erro",   "",        ok=True)
            print("  [%d] %s — OK" % (i + 1, codigo))

        except Exception as e:
            msg = repr(e).split("(")[0] + ": " + str(e).split("\n")[0][:150]
            marcar(linha, "status", "erro", ok=False)
            marcar(linha, "erro",   msg,    ok=False)
            print("  [%d] %s — ERRO: %s" % (i + 1, codigo, msg))
            try:
                esperar_clicavel(driver, By.ID, "btn-limpar-produto").click()
            except Exception:
                pass

        finally:
            wb.save(EXCEL_PATH)
            time.sleep(0.5)


def registrar_vendas(driver, wb):
    print("\n-- Vendas --")
    _, linhas = ler_aba(wb, "Vendas")
    clicar_aba(driver, "vendas")
    time.sleep(1)

    for i, linha in enumerate(linhas):
        cliente_email  = val(linha, "cliente_email")
        produto_codigo = val(linha, "produto_codigo")

        if val(linha, "status") == "sucesso":
            print("  [%d] %s / %s — pulando" % (i + 1, cliente_email, produto_codigo))
            continue

        try:
            # busca os IDs reais via API pra não depender de estado na tela
            clientes = buscar_api(driver, "/api/clientes")
            cliente  = next((c for c in clientes if c["email"] == cliente_email), None)
            if not cliente:
                raise ValueError("Cliente nao encontrado: " + cliente_email)

            produtos = buscar_api(driver, "/api/produtos")
            produto  = next((p for p in produtos if p["codigo"] == produto_codigo), None)
            if not produto:
                raise ValueError("Produto nao encontrado: " + produto_codigo)

            # registra direto via API e salva o ID real gerado pelo banco
            resp = requests.post(
                BASE_URL + "/api/vendas",
                json={
                    "cliente_id": cliente["id"],
                    "produto_id": produto["id"],
                    "quantidade": int(val(linha, "quantidade"))
                },
                timeout=10
            )
            data = resp.json()
            if not resp.ok:
                raise ValueError("API retornou erro: " + data.get("erro", str(resp.status_code)))

            pedido_id = data["id"]
            marcar(linha, "pedido_id", pedido_id, ok=True)
            marcar(linha, "status",    "sucesso",  ok=True)
            marcar(linha, "erro",      "",         ok=True)
            print("  [%d] %s / %s — OK (pedido_id: %s)" % (i + 1, cliente_email, produto_codigo, pedido_id))

        except Exception as e:
            msg = repr(e).split("(")[0] + ": " + str(e).split("\n")[0][:150]
            marcar(linha, "status", "erro", ok=False)
            marcar(linha, "erro",   msg,    ok=False)
            print("  [%d] %s / %s — ERRO: %s" % (i + 1, cliente_email, produto_codigo, msg))
            try:
                for btn in driver.find_elements(By.CSS_SELECTOR, ".btn-remove-item"):
                    btn.click()
                    time.sleep(0.1)
            except Exception:
                pass

        finally:
            wb.save(EXCEL_PATH)
            time.sleep(0.8)


def main():
    print("Automacao iniciada — %s" % datetime.now().strftime("%d/%m/%Y %H:%M:%S"))
    print("Arquivo: %s" % EXCEL_PATH)
    print("Servidor: %s" % BASE_URL)

    wb     = openpyxl.load_workbook(EXCEL_PATH)
    driver = iniciar_driver()

    try:
        esperar(driver, By.CSS_SELECTOR, ".topbar")
        time.sleep(1)

        cadastrar_clientes(driver, wb)
        cadastrar_produtos(driver, wb)
        registrar_vendas(driver, wb)

        print("\nConcluido — resultados salvos em %s" % EXCEL_PATH)

    except Exception as e:
        print("\nErro critico: %s" % str(e))
        raise

    finally:
        time.sleep(2)
        driver.quit()


if __name__ == "__main__":
    main()
